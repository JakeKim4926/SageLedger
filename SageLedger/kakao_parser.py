from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook

from .decryptor import ensure_readable_excel
from .models import KakaoTransaction

HEADER_REQUIRED = {"거래일시", "구분", "거래금액", "거래 후 잔액", "거래구분", "내용"}


def _to_int(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", "")
    text = re.sub(r"[^0-9\-]", "", text)
    return int(text or 0)


def _to_datetime(value) -> datetime:
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y.%m.%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y.%m.%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"거래일시 형식을 해석할 수 없습니다: {value!r}")


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, ws.max_row + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        normalized = {str(v).strip(): idx + 1 for idx, v in enumerate(values) if v is not None and str(v).strip()}
        if HEADER_REQUIRED.issubset(set(normalized.keys())):
            return row_idx, normalized
    raise ValueError("카카오뱅크 거래내역 헤더를 찾지 못했습니다.")


def parse_kakaobank_transactions(
    file_path: str | Path,
    password: Optional[str] = None,
    include_year_months: Optional[set[str]] = None,
) -> list[KakaoTransaction]:
    """Parse KakaoBank transaction xlsx into normalized transactions.

    include_year_months example: {"2026-06"}. If None, all rows in the file are parsed.
    """
    readable_path = ensure_readable_excel(file_path, password=password, suffix=".xlsx")
    wb = load_workbook(readable_path, read_only=True, data_only=True)
    ws = wb.active

    header_row, col_map = _find_header_row(ws)
    transactions: list[KakaoTransaction] = []

    for row_no in range(header_row + 1, ws.max_row + 1):
        raw_date = ws.cell(row_no, col_map["거래일시"]).value
        if raw_date in (None, ""):
            continue

        traded_at = _to_datetime(raw_date)
        year_month = f"{traded_at.year:04d}-{traded_at.month:02d}"
        if include_year_months and year_month not in include_year_months:
            continue

        amount = _to_int(ws.cell(row_no, col_map["거래금액"]).value)
        direction = str(ws.cell(row_no, col_map["구분"]).value or "").strip()
        balance = _to_int(ws.cell(row_no, col_map["거래 후 잔액"]).value)
        transaction_type = str(ws.cell(row_no, col_map["거래구분"]).value or "").strip()
        description = str(ws.cell(row_no, col_map["내용"]).value or "").strip()
        memo_col = col_map.get("메모")
        memo = str(ws.cell(row_no, memo_col).value or "").strip() if memo_col else ""

        transactions.append(
            KakaoTransaction(
                row_no=row_no,
                traded_at=traded_at,
                direction=direction,
                amount=amount,
                balance=balance,
                transaction_type=transaction_type,
                description=description,
                memo=memo,
            )
        )

    wb.close()
    transactions.sort(key=lambda item: item.traded_at)
    return transactions
