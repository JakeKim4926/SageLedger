from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook

from .models import Member


def _parse_bool(value) -> bool:
    if value is None:
        return True
    text = str(value).strip().lower()
    return text not in {"false", "0", "n", "no", "비활성", "탈퇴"}


def _parse_int(value, default: int = 20_000) -> int:
    if value is None or value == "":
        return default
    return int(str(value).replace(",", "").strip())


def _parse_aliases(value) -> tuple[str, ...]:
    if value is None:
        return tuple()
    parts = [part.strip() for part in str(value).replace(";", ",").split(",")]
    return tuple(part for part in parts if part)


def load_members(path: str | Path) -> list[Member]:
    """Load members from .csv or .xlsx.

    Required column: name 또는 이름
    Optional columns: monthly_due/월회비, status/상태, aliases/별칭, active/활성여부
    """
    src = Path(path)
    if not src.exists():
        raise FileNotFoundError(f"회원명단 파일을 찾을 수 없습니다: {src}")

    if src.suffix.lower() == ".csv":
        with src.open("r", encoding="utf-8-sig", newline="") as fp:
            rows = list(csv.DictReader(fp))
    elif src.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = load_workbook(src, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        wb.close()
    else:
        raise ValueError("회원명단은 csv, xlsx, xlsm 중 하나여야 합니다.")

    members: list[Member] = []
    for row in rows:
        name = row.get("name") or row.get("이름") or row.get("회원명")
        if not name:
            continue
        members.append(
            Member(
                name=str(name).strip(),
                monthly_due=_parse_int(row.get("monthly_due") or row.get("월회비"), 20_000),
                status=str(row.get("status") or row.get("상태") or "취직").strip(),
                aliases=_parse_aliases(row.get("aliases") or row.get("별칭") or ""),
                active=_parse_bool(row.get("active") or row.get("활성여부")),
            )
        )
    return [member for member in members if member.active]


def members_from_ledger_personal_sheet(
    ledger_path: str | Path,
    personal_sheet_name: str = "개인 입금 내역",
    member_header_row: int = 15,
    default_monthly_due: int = 20_000,
) -> list[Member]:
    wb = load_workbook(ledger_path, read_only=True, data_only=True, keep_vba=Path(ledger_path).suffix.lower() == ".xlsm")
    ws = wb[personal_sheet_name]
    members: list[Member] = []
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(member_header_row, col_idx).value
        next_value = ws.cell(member_header_row + 1, col_idx).value
        if value and next_value == "입금일" and str(value).strip() != "합계":
            members.append(Member(name=str(value).strip(), monthly_due=default_monthly_due))
    wb.close()
    return members
