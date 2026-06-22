from __future__ import annotations

from datetime import datetime
from typing import Iterable

from openpyxl.styles import Font, PatternFill, Alignment

from .models import LedgerWriteResult, MatchResult

REVIEW_SHEET_NAME = "자동화_검토결과"


def write_review_sheet(wb, matches: Iterable[MatchResult], result: LedgerWriteResult) -> int:
    if REVIEW_SHEET_NAME in wb.sheetnames:
        del wb[REVIEW_SHEET_NAME]
    ws = wb.create_sheet(REVIEW_SHEET_NAME)

    title = "장부 자동화 검토 결과"
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"생성시각: {datetime.now():%Y-%m-%d %H:%M:%S}"

    headers = ["거래일시", "입출금", "거래구분", "내용", "금액", "추출이름", "매칭회원", "상태", "사유", "자동반영"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(4, col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")

    row_idx = 5
    review_count = 0
    for match in matches:
        if match.auto_confirmed:
            continue
        tx = match.transaction
        row = [
            tx.traded_at,
            tx.direction,
            tx.transaction_type,
            tx.description,
            tx.amount,
            match.extracted_name,
            match.member.name if match.member else "",
            match.status.value,
            match.reason,
            "Y" if match.auto_confirmed else "N",
        ]
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx).value = value
        row_idx += 1
        review_count += 1

    summary_start = row_idx + 2
    ws.cell(summary_start, 1).value = "처리 요약"
    ws.cell(summary_start, 1).font = Font(bold=True)
    summary_rows = [
        ("개인 입금 내역 반영", result.written_personal_rows),
        ("입출금 내역 반영", result.written_ledger_rows),
        ("검토 필요", review_count),
        ("카카오뱅크 최종 잔액", result.final_bank_balance),
        ("장부 최종 잔액", result.final_ledger_balance),
    ]
    for offset, (label, value) in enumerate(summary_rows, start=1):
        ws.cell(summary_start + offset, 1).value = label
        ws.cell(summary_start + offset, 2).value = value

    if result.validation_messages:
        msg_row = summary_start + len(summary_rows) + 2
        ws.cell(msg_row, 1).value = "검증 메시지"
        ws.cell(msg_row, 1).font = Font(bold=True)
        for i, msg in enumerate(result.validation_messages, start=1):
            ws.cell(msg_row + i, 1).value = msg

    widths = [20, 10, 14, 30, 14, 12, 12, 14, 45, 10]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = "A5"
    return review_count
