from __future__ import annotations

from collections import defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .decryptor import ensure_readable_excel
from .models import KakaoTransaction, LedgerWriteResult, MatchResult


def _month_label(dt: datetime) -> str:
    return f"{dt.year}년 {dt.month}월"


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col_idx)
        dst = ws.cell(target_row, col_idx)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)


def _numeric_cell_value(value) -> Optional[int]:
    """Best-effort numeric conversion for existing ledger cells.

    The original workbook contains amount formulas such as =(4+2+2)*10000.
    openpyxl does not calculate formulas, so we parse the simple formula shape used by this ledger.
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if text.startswith("="):
        body = text[1:].replace(" ", "")
        # The ledger amount formulas are simple arithmetic only.
        # Safely evaluate numbers and + - * / parentheses, but no names/functions.
        import ast
        import operator as op

        operators = {
            ast.Add: op.add,
            ast.Sub: op.sub,
            ast.Mult: op.mul,
            ast.Div: op.truediv,
            ast.USub: op.neg,
            ast.UAdd: op.pos,
        }

        def eval_node(node):
            if isinstance(node, ast.Expression):
                return eval_node(node.body)
            if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
                return node.value
            if isinstance(node, ast.BinOp) and type(node.op) in operators:
                return operators[type(node.op)](eval_node(node.left), eval_node(node.right))
            if isinstance(node, ast.UnaryOp) and type(node.op) in operators:
                return operators[type(node.op)](eval_node(node.operand))
            raise ValueError

        try:
            return int(round(eval_node(ast.parse(body, mode="eval"))))
        except Exception:
            return None
    try:
        return int(text.replace(",", ""))
    except ValueError:
        return None


def _find_month_row(ws: Worksheet, month_col: int, start_row: int, label: str) -> Optional[int]:
    for row_idx in range(start_row, ws.max_row + 1):
        value = ws.cell(row_idx, month_col).value
        if value is None:
            continue
        if str(value).strip() == label:
            return row_idx
    return None


def _find_member_columns(ws: Worksheet, member_header_row: int) -> dict[str, tuple[int, int]]:
    columns: dict[str, tuple[int, int]] = {}
    for col_idx in range(1, ws.max_column + 1):
        name = ws.cell(member_header_row, col_idx).value
        label = ws.cell(member_header_row + 1, col_idx).value
        amount_label = ws.cell(member_header_row + 1, col_idx + 1).value if col_idx + 1 <= ws.max_column else None
        if name and label == "입금일" and amount_label == "입금액" and str(name).strip() != "합계":
            columns[str(name).strip()] = (col_idx, col_idx + 1)
    return columns


def _find_next_transaction_row(ws: Worksheet, start_row: int, date_col: int, amount_col: int) -> int:
    for row_idx in range(start_row, ws.max_row + 1):
        if ws.cell(row_idx, date_col).value is None and ws.cell(row_idx, amount_col).value is None:
            return row_idx
    ws.insert_rows(ws.max_row + 1)
    _copy_row_style(ws, ws.max_row - 1, ws.max_row)
    return ws.max_row


def _last_balance_before(ws: Worksheet, row_idx: int, balance_col: int) -> int:
    # Rebuild balance from the ledger rows. This is more reliable than reading the W column
    # because most existing W cells are formulas and openpyxl does not calculate them.
    type_col = 5
    amount_col = 19
    balance = 0
    touched = False
    for cursor in range(17, row_idx):
        kind = ws.cell(cursor, type_col).value
        amount = _numeric_cell_value(ws.cell(cursor, amount_col).value)
        if kind in ("수입", "지출") and amount is not None:
            balance = balance + amount if kind == "수입" else balance - amount
            touched = True
    if touched:
        return balance

    # Fallback for non-standard sheets.
    for cursor in range(row_idx - 1, 0, -1):
        direct = _numeric_cell_value(ws.cell(cursor, balance_col).value)
        if direct is not None:
            return direct
    return 0


def _detail_exists(ws: Worksheet, detail_col: int, amount_col: int, detail: str, amount: int) -> bool:
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row_idx, detail_col).value == detail and _numeric_cell_value(ws.cell(row_idx, amount_col).value) == amount:
            return True
    return False


def _transaction_exists(ws: Worksheet, cfg: dict, traded_at: datetime, kind: str, summary: str, amount: int) -> bool:
    date_col = int(cfg["columns"]["date"])
    type_col = int(cfg["columns"]["type"])
    summary_col = int(cfg["columns"]["summary"])
    amount_col = int(cfg["columns"]["amount"])
    for row_idx in range(int(cfg["start_row"]), ws.max_row + 1):
        row_date = ws.cell(row_idx, date_col).value
        if hasattr(row_date, "date"):
            row_date = row_date.date()
        elif isinstance(row_date, str):
            try:
                row_date = datetime.fromisoformat(row_date).date()
            except ValueError:
                pass
        if (
            row_date == traded_at.date()
            and ws.cell(row_idx, type_col).value == kind
            and ws.cell(row_idx, summary_col).value == summary
            and _numeric_cell_value(ws.cell(row_idx, amount_col).value) == amount
        ):
            return True
    return False


def load_ledger_workbook(ledger_path: str | Path, password: Optional[str] = None):
    readable_path = ensure_readable_excel(ledger_path, password=password, suffix=Path(ledger_path).suffix)
    return load_workbook(readable_path, keep_vba=Path(ledger_path).suffix.lower() == ".xlsm")


def write_personal_deposits(wb, matches: Iterable[MatchResult], mapping: dict, write_review_required: bool = False) -> tuple[int, set[int]]:
    cfg = mapping["personal_deposit_sheet"]
    ws = wb[cfg["name"]]
    member_cols = _find_member_columns(ws, int(cfg["member_header_row"]))
    written = 0
    written_tx_rows: set[int] = set()

    for result in matches:
        if not result.member:
            continue
        if not result.auto_confirmed and not write_review_required:
            continue

        tx = result.transaction
        label = _month_label(tx.traded_at)
        row_idx = _find_month_row(ws, int(cfg["month_col"]), int(cfg["start_month_row"]), label)
        if row_idx is None:
            continue

        columns = member_cols.get(result.member.name)
        if not columns:
            continue
        date_col, amount_col = columns

        existing_amount = ws.cell(row_idx, amount_col).value
        numeric_existing = _numeric_cell_value(existing_amount)
        if numeric_existing is not None and numeric_existing == tx.amount:
            continue
        if numeric_existing is not None and not write_review_required:
            continue

        ws.cell(row_idx, date_col).value = tx.traded_at.date()
        ws.cell(row_idx, amount_col).value = tx.amount
        written += 1
        written_tx_rows.add(tx.row_no)

    return written, written_tx_rows


def _append_transaction_row(ws: Worksheet, cfg: dict, traded_at: datetime, kind: str, summary: str, detail: str, amount: int) -> bool:
    start_row = int(cfg["start_row"])
    seq_col = int(cfg["columns"]["seq"])
    date_col = int(cfg["columns"]["date"])
    type_col = int(cfg["columns"]["type"])
    summary_col = int(cfg["columns"]["summary"])
    detail_col = int(cfg["columns"]["detail"])
    amount_col = int(cfg["columns"]["amount"])
    balance_col = int(cfg["columns"]["balance"])

    if _detail_exists(ws, detail_col, amount_col, detail, amount):
        return False

    row_idx = _find_next_transaction_row(ws, start_row, date_col, amount_col)
    if row_idx > start_row:
        _copy_row_style(ws, row_idx - 1, row_idx)

    prev_seq = ws.cell(row_idx - 1, seq_col).value if row_idx > start_row else None
    if isinstance(prev_seq, (int, float)):
        seq = int(prev_seq) + 1
    else:
        seq = row_idx - start_row + 1

    prev_balance = _last_balance_before(ws, row_idx, balance_col)
    next_balance = prev_balance + amount if kind == "수입" else prev_balance - amount

    ws.cell(row_idx, seq_col).value = seq
    ws.cell(row_idx, date_col).value = traded_at.date()
    ws.cell(row_idx, type_col).value = kind
    ws.cell(row_idx, summary_col).value = summary
    ws.cell(row_idx, detail_col).value = detail
    ws.cell(row_idx, amount_col).value = amount
    ws.cell(row_idx, balance_col).value = next_balance
    return True


def write_transaction_ledger(
    wb,
    transactions: Iterable[KakaoTransaction],
    matches: Iterable[MatchResult],
    mapping: dict,
    category_rules: dict,
    written_transaction_rows: set[int] | None = None,
    write_review_required: bool = False,
) -> int:
    cfg = mapping["transaction_sheet"]
    ws = wb[cfg["name"]]
    written = 0

    confirmed_income_by_month: dict[str, int] = defaultdict(int)
    confirmed_date_by_month: dict[str, datetime] = {}
    written_transaction_rows = written_transaction_rows or set()
    for result in matches:
        # 입출금 내역에는 이번 실행에서 개인 입금 내역에 실제로 새로 쓴 입금만 회비 정산으로 반영한다.
        # 이렇게 해야 이미 장부에 들어간 과거 입금을 다시 합산하지 않는다.
        if result.transaction.row_no in written_transaction_rows:
            tx = result.transaction
            key = tx.year_month_key
            confirmed_income_by_month[key] += tx.amount
            confirmed_date_by_month[key] = max(confirmed_date_by_month.get(key, tx.traded_at), tx.traded_at)

    for key, amount in sorted(confirmed_income_by_month.items()):
        year, month = key.split("-")
        dt = confirmed_date_by_month[key]
        detail = f"자동화 {int(year)}년 {int(month)}월 회비 정산"
        if _append_transaction_row(ws, cfg, dt, "수입", "회비 정산", detail, amount):
            written += 1

    for tx in sorted(transactions, key=lambda item: item.traded_at):
        if tx.transaction_type == "예금이자" and tx.income > 0:
            if _transaction_exists(ws, cfg, tx.traded_at, "수입", "이자", tx.income):
                continue
            detail = f"자동화 카카오뱅크 예금이자 {tx.traded_at:%Y-%m-%d}"
            if _append_transaction_row(ws, cfg, tx.traded_at, "수입", "이자", detail, tx.income):
                written += 1
        elif tx.expense > 0:
            summary = categorize_expense(tx.description, category_rules)
            if _transaction_exists(ws, cfg, tx.traded_at, "지출", summary, tx.expense):
                continue
            detail = f"자동화 {tx.description}" if tx.description else f"자동화 출금 {tx.traded_at:%Y-%m-%d}"
            if _append_transaction_row(ws, cfg, tx.traded_at, "지출", summary, detail, tx.expense):
                written += 1

    return written


def categorize_expense(description: str, category_rules: dict) -> str:
    default_category = category_rules.get("default_expense_category", "기타 지출")
    rules = category_rules.get("expense_keywords", {})
    text = description or ""
    for keyword, category in rules.items():
        if keyword in text:
            return str(category)
    return default_category


def save_workbook(wb, output_path: str | Path) -> Path:
    dst = Path(output_path)
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    return dst
