from __future__ import annotations

import ast
import operator as op
import re
from collections import defaultdict
from datetime import date, datetime
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .config import GroupConfig
from .decryptor import ensure_readable_excel
from .models import CellWrite, KakaoTransaction, ReviewItem, WritePlan
from .name_matcher import match_member_deposits

_OPS = {ast.Add: op.add, ast.Sub: op.sub, ast.Mult: op.mul, ast.Div: op.truediv,
        ast.USub: op.neg, ast.UAdd: op.pos}


def _eval_formula_number(text: str) -> Optional[int]:
    """=(2+2)*10000 같은 단순 산술 수식만 안전하게 계산한다 (장부 금액 수식 형태)."""
    body = text[1:].replace(" ", "")

    def ev(node):
        if isinstance(node, ast.Expression):
            return ev(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        if isinstance(node, ast.BinOp) and type(node.op) in _OPS:
            return _OPS[type(node.op)](ev(node.left), ev(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _OPS:
            return _OPS[type(node.op)](ev(node.operand))
        raise ValueError
    try:
        return int(round(ev(ast.parse(body, mode="eval"))))
    except Exception:
        return None


def _numeric(value) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if text.startswith("="):
        return _eval_formula_number(text)
    try:
        return int(text.replace(",", ""))
    except ValueError:
        return None


def _as_date(value):
    if hasattr(value, "date"):
        return value.date()
    return value


def categorize_expense(description: str, category_rules: dict) -> str:
    default = category_rules.get("default_expense_category", "기타 지출")
    for keyword, category in (category_rules.get("expense_keywords") or {}).items():
        if keyword in (description or ""):
            return str(category)
    return default


def build_plan(
    group: GroupConfig,
    transactions: list[KakaoTransaction],
    mapping: dict,
    category_rules: dict,
) -> WritePlan:
    readable = ensure_readable_excel(group.ledger_path, password=group.read_pw,
                                     suffix=group.ledger_path.suffix)
    wb = load_workbook(readable, data_only=False,
                       keep_vba=group.ledger_path.suffix.lower() == ".xlsm")
    plan = WritePlan()

    matched, unmatched = match_member_deposits(transactions, group.members)
    for tx, reason in unmatched:
        plan.reviews.append(ReviewItem(tx.traded_at, "(미매칭)", tx.amount, reason))

    deposits_by_member: dict[str, list[KakaoTransaction]] = defaultdict(list)
    for member, tx in matched:
        deposits_by_member[member.name].append(tx)
    due_by_member = {m.name: m.monthly_due for m in group.members}

    # 회비정산용: 실제로 새로 기입한 개인입금을 입금월별로 집계.
    monthly_written: dict[str, list] = {}  # ym -> [amount_sum, max_date]

    memo_incomes: list[tuple[str, KakaoTransaction]] = []
    _plan_personal(wb, mapping, deposits_by_member, due_by_member,
                   monthly_written, plan, memo_incomes)

    # 메모 'X' 로 표시된 비회비 입금: 세부내역을 입력받아 기타 수입으로 기록.
    extra_incomes: list[tuple] = []  # (date, "수입", "기타", detail, amount)
    for member_name, tx in memo_incomes:
        print(f"\n[기타 수입] {member_name} {tx.amount:,}원 (메모 X) - 회비 아닌 입금.")
        detail = input("  세부내역 입력 (빈칸=검토로 남김): ").strip()
        if detail:
            extra_incomes.append((tx.traded_at.date(), "수입", "기타", detail, tx.amount))
        else:
            plan.reviews.append(ReviewItem(tx.traded_at, member_name, tx.amount,
                                "회비 아닌 입금 (세부내역 미입력)"))

    _plan_transactions(wb, mapping, transactions, category_rules,
                       monthly_written, plan, extra_incomes)

    plan.bank_final_balance = transactions[-1].balance if transactions else None
    wb.close()
    return plan


def _plan_personal(wb, mapping, deposits_by_member, due_by_member,
                   monthly_written, plan: WritePlan, memo_incomes: list) -> None:
    cfg = mapping["personal_deposit_sheet"]
    name = cfg["name"]
    ws = wb[name]
    header_row = int(cfg["member_header_row"])
    month_col = int(cfg["month_col"])
    start_row = int(cfg["start_month_row"])

    # 회원 -> (입금일 열, 입금액 열)
    member_cols: dict[str, tuple[int, int]] = {}
    for c in range(1, ws.max_column + 1):
        nm = ws.cell(header_row, c).value
        lab = ws.cell(header_row + 1, c).value
        lab2 = ws.cell(header_row + 1, c + 1).value if c + 1 <= ws.max_column else None
        if nm and str(nm).strip() != "합계" and lab == "입금일" and lab2 == "입금액":
            member_cols[str(nm).strip()] = (c, c + 1)

    # '총계' 경계 행 (없으면 시트 끝)
    boundary = ws.max_row + 1
    for r in range(start_row, ws.max_row + 1):
        if str(ws.cell(r, month_col).value).strip() == "총계":
            boundary = r
            break

    for member_name, txs in deposits_by_member.items():
        cols = member_cols.get(member_name)
        if not cols:
            plan.reviews.append(ReviewItem(None, member_name, 0,
                                "개인 입금 내역 시트에 회원 열이 없음"))
            continue
        date_col, amt_col = cols
        due = due_by_member[member_name]

        existing: set[tuple] = set()
        empty_rows: list[int] = []
        for r in range(start_row, boundary):
            dval = ws.cell(r, date_col).value
            aval = ws.cell(r, amt_col).value
            if dval is None and aval is None:
                empty_rows.append(r)
            else:
                existing.add((_as_date(dval), _numeric(aval)))

        # 입금액을 월회비 단위로 등분. 배수가 아니면 검토.
        chunks: list[tuple[date, int]] = []
        for tx in sorted(txs, key=lambda t: t.traded_at):
            if tx.amount % due != 0:
                if str(tx.memo).strip().upper() == "X":
                    memo_incomes.append((member_name, tx))
                else:
                    plan.reviews.append(ReviewItem(tx.traded_at, member_name, tx.amount,
                                        f"월회비 {due:,}원의 배수가 아님 (등분 불가)"))
                continue
            for _ in range(tx.amount // due):
                chunks.append((tx.traded_at.date(), due))

        # 이미 장부에 있는 (날짜,금액) 은 건너뜀.
        to_write = [(d, a) for (d, a) in chunks if (d, a) not in existing]

        if len(to_write) > len(empty_rows):
            plan.notes.append(
                f"{member_name}: 빈칸 부족 (필요 {len(to_write)}칸 / 남은 {len(empty_rows)}칸). "
                f"넘치는 {len(to_write) - len(empty_rows)}건은 검토로 분리.")
            for d, a in to_write[len(empty_rows):]:
                plan.reviews.append(ReviewItem(datetime(d.year, d.month, d.day),
                                    member_name, a, "개인 입금 내역 빈칸 부족"))
            to_write = to_write[:len(empty_rows)]

        for (d, a), row in zip(to_write, empty_rows):
            dt = datetime(d.year, d.month, d.day)
            plan.cells.append(CellWrite(name, row, date_col, dt))
            plan.cells.append(CellWrite(name, row, amt_col, a))
            plan.personal_written += 1
            ym = f"{d.year:04d}-{d.month:02d}"
            bucket = monthly_written.setdefault(ym, [0, dt])
            bucket[0] += a
            bucket[1] = max(bucket[1], dt)


def _plan_transactions(wb, mapping, transactions, category_rules,
                       monthly_written, plan: WritePlan, extra_incomes: list) -> None:
    cfg = mapping["transaction_sheet"]
    name = cfg["name"]
    ws = wb[name]
    cols = {k: int(v) for k, v in cfg["columns"].items()}
    start = int(cfg["start_row"])

    last = start - 1
    for r in range(start, ws.max_row + 1):
        if ws.cell(r, cols["date"]).value is not None:
            last = r

    existing: set[tuple] = set()
    for r in range(start, last + 1):
        existing.add((
            _as_date(ws.cell(r, cols["date"]).value),
            ws.cell(r, cols["type"]).value,
            ws.cell(r, cols["summary"]).value,
            _numeric(ws.cell(r, cols["amount"]).value),
        ))

    prev_seq = ws.cell(last, cols["seq"]).value if last >= start else 0

    new_entries: list[tuple] = []  # (date, 구분, 내역, 세부, 금액)

    # 1) 회비 정산 (이번에 새로 채운 개인입금을 입금월별로)
    for ym, (amount, maxd) in sorted(monthly_written.items()):
        y, m = ym.split("-")
        detail = f"자동화 {int(y)}년 {int(m)}월 회비 정산"
        key = (maxd.date(), "수입", "회비 정산", amount)
        if key not in existing:
            new_entries.append((maxd.date(), "수입", "회비 정산", detail, amount))

    # 2) 예금이자
    for tx in sorted(transactions, key=lambda t: t.traded_at):
        if tx.transaction_type == "예금이자" and tx.income > 0:
            key = (tx.traded_at.date(), "수입", "이자", tx.income)
            if key not in existing:
                new_entries.append((tx.traded_at.date(), "수입", "이자",
                                    f"자동화 카카오뱅크 예금이자 {tx.traded_at:%Y-%m-%d}", tx.income))

    # 3) 지출 (출금)
    for tx in sorted(transactions, key=lambda t: t.traded_at):
        if tx.expense > 0:
            summary = categorize_expense(tx.description, category_rules)
            key = (tx.traded_at.date(), "지출", summary, tx.expense)
            if key not in existing:
                detail = f"자동화 {tx.description}" if tx.description else \
                         f"자동화 출금 {tx.traded_at:%Y-%m-%d}"
                new_entries.append((tx.traded_at.date(), "지출", summary, detail, tx.expense))

    # 4) 기타 수입 (메모 'X' 로 표시된 비회비 입금)
    for d, kind, summary, detail, amount in extra_incomes:
        if (d, kind, summary, amount) not in existing:
            new_entries.append((d, kind, summary, detail, amount))

    new_entries.sort(key=lambda e: e[0])

    el = get_column_letter(cols["type"])
    sl = get_column_letter(cols["amount"])
    wl = get_column_letter(cols["balance"])

    row = last
    seq = int(prev_seq) if isinstance(prev_seq, (int, float)) else last - start
    for d, kind, summary, detail, amount in new_entries:
        row += 1
        seq += 1
        dt = datetime(d.year, d.month, d.day)
        plan.cells.append(CellWrite(name, row, cols["seq"], seq))
        plan.cells.append(CellWrite(name, row, cols["date"], dt))
        plan.cells.append(CellWrite(name, row, cols["type"], kind))
        plan.cells.append(CellWrite(name, row, cols["summary"], summary))
        plan.cells.append(CellWrite(name, row, cols["detail"], detail))
        plan.cells.append(CellWrite(name, row, cols["amount"], amount))
        plan.cells.append(CellWrite(name, row, cols["balance"],
                          f'=IF({el}{row}="수입",{wl}{row-1}+{sl}{row},{wl}{row-1}-{sl}{row})',
                          formula=True))
        plan.transaction_written += 1

    last_new_row = row

    # 수입/지출 합계 SUMIF 고정범위 확장 (필요 시)
    for crow in (15, 16):
        f = ws.cell(crow, 30).value  # AD열
        if isinstance(f, str) and f.startswith("=") and last_new_row > 215:
            new_f = re.sub(r"(\$[A-Z]+\$17:\$[A-Z]+\$)\d+",
                           lambda mm: mm.group(1) + str(last_new_row), f)
            if new_f != f:
                plan.cells.append(CellWrite(name, crow, 30, new_f, formula=True))

    plan.last_transaction_row = last_new_row
    plan.balance_col = cols["balance"]
